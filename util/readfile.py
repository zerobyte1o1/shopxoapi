import openpyxl

#打开Excel表格
def readexecl(filename,sheetname):
    wb = openpyxl.load_workbook(filename)

    #指定要读取内容的sheet
    sheetobj = wb[sheetname]

    #读取所有的内容

    #获取最大行
    rows = sheetobj.max_row

    #获取最大列
    columns = sheetobj.max_column

    list = []

    # 使用for循环，
    for row in range(2, rows+1):
        #定义一个列表，用来存放我们获取的每一个数据
        new_list = []
        #列的处理
        for column in range(1, columns+1):
            #取值
            new_list.append(sheetobj.cell(row,column).value)

        list.append(new_list)

    return list